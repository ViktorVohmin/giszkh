#!/usr/bin/python
# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
import pathlib
from openpyxl.styles import Color, PatternFill, Font, Border
import time
from datetime import datetime
import os
from inspect import getsourcefile
from os.path import abspath
from os import path
import PySimpleGUI as sg
import openpyxl

url="https://my.dom.gosuslugi.ru/organization-cabinet/#!/debts/received-requests"

def bclick(driver, link):
    while True:
        if len(driver.find_elements_by_css_selector(link)) > 0:
            driver.find_element_by_css_selector(link).click()
            break;

def findstr(driver, name):
    xpath = "//*[contains(text(), '" + str(name) + "')]" # simplified
    list = driver.find_elements_by_xpath(xpath) # locate all elements by xpath
    if len(list) > 0: # if list is not empty, click on element
        list[0].click() # click on the first element in the list
    time.sleep(5)
    
def send(driver, pfile, rfile, application_path):
    # читаем excel-файл
    try:
        s = []
        start_time = datetime.now()
    
        rwb = openpyxl.load_workbook(rfile)
        rsheet = rwb.worksheets[0]
        rmax_row = rsheet.max_row
        
        adr1=''
        adr2=''
        
        wb = openpyxl.load_workbook(pfile)
        sheet = wb.worksheets[0]
        max_row = sheet.max_row
        min_row = sheet.min_row   
        
        if path.isfile(application_path+'\\Отчет.xlsx'):
            report = openpyxl.load_workbook(application_path+'\\Отчет.xlsx')
            repsheet = report.worksheets[0]
            repmax_row = repsheet.max_row
            for row_num in range(1, repmax_row):
                if (str(repsheet.cell(row = row_num+1, column =20).value)=='None') and (str(sheet.cell(row = row_num+1, column = 2).value)==str(repsheet.cell(row = row_num+1, column =1).value)):
                    min_row = repsheet.cell(row = row_num+1, column =20).row
                    break      
        for rrow_num in range(1, rmax_row): 
            if (min_row>1):
                star=min_row
            else:
                star=min_row+1
            for row_num in range(star, max_row):
                if str(rsheet.cell(row = rrow_num+1, column = 1).value)!='None':
                    adr1=str(rsheet.cell(row = rrow_num+1, column = 1).value)+'|'+str(rsheet.cell(row = rrow_num+1, column = 2).value)+'|'+str(rsheet.cell(row = rrow_num+1, column = 3).value)
                if str(sheet.cell(row = row_num+1, column = 2).value)!='None':
                    adr2=str(sheet.cell(row = row_num+1, column = 9).value)+'|'+str(sheet.cell(row = row_num+1, column = 10).value)+'|'+str(sheet.cell(row = row_num+1, column = 11).value)
                if adr1==adr2:
                    time.sleep(5)
                    driver.find_element_by_css_selector("body > div.page-wrapper > div.app-content-wrapper > div > div > div:nth-child(1) > div > debtreq-received-requests-search > ng-form > div > ef-bp-form > div > form > div.collapse-toggle__cnt.collapse.in > div > div > div > div > div:nth-child(1) > div.col-xs-7 > div:nth-child(1) > div > div > input").send_keys(Keys.CONTROL, 'a')
                    driver.find_element_by_css_selector("body > div.page-wrapper > div.app-content-wrapper > div > div > div:nth-child(1) > div > debtreq-received-requests-search > ng-form > div > ef-bp-form > div > form > div.collapse-toggle__cnt.collapse.in > div > div > div > div > div:nth-child(1) > div.col-xs-7 > div:nth-child(1) > div > div > input").send_keys(Keys.BACKSPACE)
                    driver.find_element_by_css_selector("body > div.page-wrapper > div.app-content-wrapper > div > div > div:nth-child(1) > div > debtreq-received-requests-search > ng-form > div > ef-bp-form > div > form > div.collapse-toggle__cnt.collapse.in > div > div > div > div > div:nth-child(1) > div.col-xs-7 > div:nth-child(1) > div > div > input").send_keys(str(sheet.cell(row = row_num+1, column = 2).value))
                    driver.find_element_by_css_selector("body > div.page-wrapper > div.app-content-wrapper > div > div > div:nth-child(1) > div > debtreq-received-requests-search > ng-form > div > ef-bp-form > div > form > div.collapse-toggle__pane.collapse-toggle__pane_utility.fix-ef-bp > div > div.col-xs-8.text-right.ng-scope > button").click()
                    time.sleep(10)
                    search=[]
                    search=driver.find_elements_by_class_name('register-card__header-title')
                    for element in search:
                        driver.get(element.get_attribute("href"))
                        break
                    time.sleep(10)
                    try:
                        driver.find_element_by_css_selector("body > div.page-wrapper > div.app-content-wrapper > div > div > div > div.section-base__footer > div > div.col-xs-6.text-right > button").click()
                        time.sleep(15)
                        try:
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.in > div > div > div > form > div:nth-child(1) > div > div.form-base__body > div > div.col-xs-4 > div:nth-child(1) > label > input").click()
                            time.sleep(10)          
                            F,I,O = str(rsheet.cell(row = rrow_num+1, column = 4).value).split()
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.in > div > div > div > form > debtreq-debt-person-info-form > ng-form > div > div.form-base__body > div.row > div:nth-child(1) > div > input").send_keys(F)
                            time.sleep(5)
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.in > div > div > div > form > debtreq-debt-person-info-form > ng-form > div > div.form-base__body > div.row > div:nth-child(2) > div > input").send_keys(I)
                            time.sleep(5)
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.in > div > div > div > form > debtreq-debt-person-info-form > ng-form > div > div.form-base__body > div.row > div:nth-child(3) > div > input").send_keys(O)
                            time.sleep(5)
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.in > div > div > div > form > debtreq-debt-person-info-form > ng-form > div > div.form-base__body > div.form-horizontal > div:nth-child(4) > div > debtreq-attachment > ef-prf-form > div > div.upload-area-wrapper.file-panel__group.without-wrap.ng-scope > div.file-panel__row > div.file-panel__row-item.file-panel__controls > label > input").send_keys(application_path+'\\out\\'+str(rsheet.cell(row=rrow_num+1, column = 5).value)+'.pdf')
                            time.sleep(5)
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.in > div > div > div > div.modal-footer.modal-base__footer > button.btn.btn-action").click()
                            time.sleep(10)
                            driver.find_element_by_css_selector("body > div.page-wrapper > div.app-content-wrapper > div > div > div > div.section-base__footer > div > div.col-xs-6.text-right > button").click()
                            time.sleep(10)
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.z-index-xxl.in > div > div > div > div.modal-footer.modal-base__footer.text-center > button.btn.btn-action").click()
                            time.sleep(10)
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.z-index-xxl.in > div > div > div > div.modal-footer.modal-base__footer.text-center > button").click()
                            time.sleep(5)
                            sheet.cell(row = row_num+1, column =20).value = 'Отработано'
                            wb.save(application_path+'\\Отчет.xlsx')
                            driver.get(url)
                            time.sleep(10)
                        except NoSuchElementException as e:
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.z-index-xxl.in > div > div > div > div.modal-footer.modal-base__footer.text-center > button.btn.btn-action").click()
                            time.sleep(10)
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.z-index-xxl.in > div > div > div > div.modal-footer.modal-base__footer.text-center > button").click()
                            time.sleep(10)
                            sheet.cell(row = row_num+1, column =20).value = 'Отработано'
                            wb.save(application_path+'\\Отчет.xlsx')
                            driver.get(url)
                            time.sleep(10)
                    except NoSuchElementException as e:
                        driver.get(url)
                        time.sleep(10)
                        
        answer = sg.popup_yes_no('Задолженности переданы, передать остальные запросы ?')
        if answer == 'No':
            #sg.popup_cancel('Файл обработан')
            sys.exit()
        else:
            if  path.isfile(application_path+'\\Отчет.xlsx'):
                report = openpyxl.load_workbook(application_path+'\\Отчет.xlsx')
                repsheet = report.worksheets[0]
                repmax_row = repsheet.max_row
            else:
                report = openpyxl.load_workbook(pfile)
                repsheet = report.worksheets[0]
                repmax_row = repsheet.max_row
            for row_num in range(1, repmax_row):
                if (str(repsheet.cell(row = row_num+1, column =20).value)!='Отработано') or (str(repsheet.cell(row = row_num+1, column =20).value)=='None'):
                    driver.get(url)
                    time.sleep(5)
                    driver.find_element_by_css_selector("body > div.page-wrapper > div.app-content-wrapper > div > div > div:nth-child(1) > div > debtreq-received-requests-search > ng-form > div > ef-bp-form > div > form > div.collapse-toggle__cnt.collapse.in > div > div > div > div > div:nth-child(1) > div.col-xs-7 > div:nth-child(1) > div > div > input").send_keys(Keys.CONTROL, 'a')
                    driver.find_element_by_css_selector("body > div.page-wrapper > div.app-content-wrapper > div > div > div:nth-child(1) > div > debtreq-received-requests-search > ng-form > div > ef-bp-form > div > form > div.collapse-toggle__cnt.collapse.in > div > div > div > div > div:nth-child(1) > div.col-xs-7 > div:nth-child(1) > div > div > input").send_keys(Keys.BACKSPACE)
                    driver.find_element_by_css_selector("body > div.page-wrapper > div.app-content-wrapper > div > div > div:nth-child(1) > div > debtreq-received-requests-search > ng-form > div > ef-bp-form > div > form > div.collapse-toggle__cnt.collapse.in > div > div > div > div > div:nth-child(1) > div.col-xs-7 > div:nth-child(1) > div > div > input").send_keys(str(repsheet.cell(row = row_num+1, column = 1).value))
                    driver.find_element_by_css_selector("body > div.page-wrapper > div.app-content-wrapper > div > div > div:nth-child(1) > div > debtreq-received-requests-search > ng-form > div > ef-bp-form > div > form > div.collapse-toggle__pane.collapse-toggle__pane_utility.fix-ef-bp > div > div.col-xs-8.text-right.ng-scope > button").click()
                    time.sleep(10)
                    search=[]
                    search=driver.find_elements_by_class_name('register-card__header-title')
                    for element in search:
                        driver.get(element.get_attribute("href"))
                        break
                    time.sleep(10)
                    try:
                        driver.find_element_by_css_selector("body > div.page-wrapper > div.app-content-wrapper > div > div > div > div.section-base__footer > div > div.col-xs-6.text-right > button").click()
                        time.sleep(15)
                        try:
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.in > div > div > div > form > div:nth-child(1) > div > div.form-base__body > div > div.col-xs-4 > div:nth-child(2) > label > input").click()
                            time.sleep(5)
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.in > div > div > div > div.modal-footer.modal-base__footer > button.btn.btn-action").click()
                            time.sleep(10)
                            driver.find_element_by_css_selector("body > div.page-wrapper > div.app-content-wrapper > div > div > div > div.section-base__footer > div > div.col-xs-6.text-right > button").click()
                            time.sleep(6)
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.z-index-xxl.in > div > div > div > div.modal-footer.modal-base__footer.text-center > button.btn.btn-action").click()
                            time.sleep(6)
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.z-index-xxl.in > div > div > div > div.modal-footer.modal-base__footer.text-center > button").click()
                            time.sleep(5)
                            repsheet.cell(row = row_num+1, column =20).value = 'Отработано'
                            report.save(application_path+'\\Отчет.xlsx')
                        except NoSuchElementException as e:
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.z-index-xxl.in > div > div > div > div.modal-footer.modal-base__footer.text-center > button.btn.btn-action").click()
                            time.sleep(10)
                            driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.z-index-xxl.in > div > div > div > div.modal-footer.modal-base__footer.text-center > button").click()
                            time.sleep(5)
                            repsheet.cell(row = row_num+1, column =20).value = 'Отработано'
                            report.save(application_path+'\\Отчет.xlsx')
                    except NoSuchElementException as e:
                        driver.get(url)
                        time.sleep(10)
    except Exception as e:
       sg.popup_cancel('Ошибка перезапустите программу') 
    finally:
        sg.popup_cancel('Файл обработан')
        if  path.isfile(application_path+'\\Отчет.xlsx'):
            os.remove(application_path+'\\Отчет.xlsx')
        
def account(PName,PPassword,POGRN,PFILE,RFILE):
    if getattr(sys, 'frozen', False):
        application_path = os.path.dirname(sys.executable)
    elif __file__:
        application_path = os.path.dirname(__file__)    
    dir = os.path.abspath(os.path.dirname(__file__))
    driver = webdriver.Chrome(executable_path=dir+'\\chromedriver.exe')
    driver.get("https://dom.gosuslugi.ru/")
    time.sleep(10)
    driver.find_element_by_css_selector("a[ng-click='sign()']").click()
    time.sleep(10)
    driver.find_element_by_css_selector("#login").send_keys(PName)
    driver.find_element_by_css_selector("#password").send_keys(PPassword)
    bclick(driver, "body > esia-root > div > esia-login > div > div.form-container > form > div.mb-24 > button")
    time.sleep(10)
    findstr(driver, POGRN)
    #bclick(driver, "#app > form > div > button")
    bclick(driver, "#saveCookie")
    bclick(driver, "#bContinue")
    time.sleep(10)
    try:
        driver.find_element_by_css_selector("body > div.modal.fade.ng-isolate-scope.in > div > div > div > div.modal-header.modal-base__header > div > button").click() 
    except NoSuchElementException:
        pass
    for element in driver.find_elements_by_css_selector('span'):
        if element.text=='Объекты управления':
            element.click()
            break
            
    for element in driver.find_elements_by_css_selector('span'):
        if element.text=='Ответ на полученные запросы о наличии задолженности за ЖКУ':
            element.click()
            break
    time.sleep(10)
    send(driver,PFILE,RFILE,application_path)

def panel():
    
    if getattr(sys, 'frozen', False):
        application_path = os.path.dirname(sys.executable)
    elif __file__:
        application_path = os.path.dirname(__file__) 
        
    if os.path.isdir('./out') is not True:
        os.makedirs('./out')
    
    layout = [[sg.Text("Подключение к ГИС ЖКХ:")],
       [sg.Text("Логин:"),sg.Input(size=(40, 1),key='-PName-')],
       [sg.Text("Пароль:"),sg.Input(size=(40, 1),key='-PPassword-', password_char='*')],
       [sg.Text("ОГРН:"),sg.Input(size=(40, 1),key='-POGRN-')],
       [sg.Text('Реестр ГИС ЖКХ:'), sg.InputText(size=(40, 1), key='-PFILE-'), sg.FileBrowse()],
       [sg.Text('Реестр судебных задолженностей:'), sg.InputText(size=(40, 1), key='-RFILE-'), sg.FileBrowse()],
       [sg.Button('Передать задолженности'), sg.Button('Отмена')],
    ]
    window = sg.Window('Гис ЖКХ',layout,icon=application_path+'\\icon.ico')

    while True:
        event, values = window.read()
        window.close()
        if event == sg.WINDOW_CLOSED or event == 'Отмена':
            break
        else:
        #print(values['-IP-'])
            try:
                account(values['-PName-'],values['-PPassword-'],values['-POGRN-'],values['-PFILE-'],values['-RFILE-'])
            except Exception as e:
                print(str(e))
                
if __name__ == '__main__':
    panel()
